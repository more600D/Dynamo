import sys
pyt_path = r'C:\Python27\Lib\site-packages'
sys.path.append(pyt_path)
import openpyxl

import clr
clr.AddReference("RevitAPI")
from Autodesk.Revit.DB import BuiltInCategory, FilteredElementCollector
clr.AddReference("RevitServices")
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager
doc = DocumentManager.Instance.CurrentDBDocument


def set_data_to_parameter(elem, param_name, data):
    document = elem.Document
    param = elem.LookupParameter(param_name)
    result = "Done"
    if param:
        param.Set(data)
    else:
        elem_type = document.GetElement(elem.GetTypeId())
        elem_type_param = elem_type.LookupParameter(param_name)
        if elem_type_param:
            elem_type_param.Set(data)
        else:
            result = "There is no such parameter name"
    return result


def check_parameter(elem, param_name):
    document = elem.Document
    info = [elem, param_name]
    result = False
    param = elem.LookupParameter(param_name)
    if param:
        value = param.AsValueString()
        if value is None or value == "":
            result = True
    else:
        elem_type = document.GetElement(elem.GetTypeId())
        elem_type_param = elem_type.LookupParameter(param_name)
        if elem_type_param:
            value = elem_type_param.AsValueString()
            if value is None or value == "":
                result = True
    return result


categories = {
    'Стены': BuiltInCategory.OST_Walls,
    'Перекрытия': BuiltInCategory.OST_Floors
}

xpath = r"C:\Users\s.shvydko\Desktop\тест1.xlsx"
workbook = openpyxl.load_workbook(xpath)
sheet = workbook['Лист1']
max_column = sheet.max_column
max_row = sheet.max_row

val = []

TransactionManager.Instance.EnsureInTransaction(doc)

error = []
result = []
for i in range(2, max_row + 1):
    cell_cat = sheet.cell(row=i, column=1).value
    try:
        cat = categories[cell_cat]
        elems = FilteredElementCollector(doc).OfCategory(cat).WhereElementIsNotElementType().ToElements()
        info = []
        for elem in elems:
            for y in range(2, max_column + 1):
                cell_param_value = sheet.cell(row=i, column=y).value
                if cell_param_value is not None:
                    cell_param_name = sheet.cell(row=1, column=y).value
                    if check_parameter(elem, cell_param_name):
                        info.append(elem)
                        info.append(cell_param_name)
        result.append(info)
    except KeyError:
        error.append(cell_cat)

TransactionManager.Instance.TransactionTaskDone()

OUT = result
