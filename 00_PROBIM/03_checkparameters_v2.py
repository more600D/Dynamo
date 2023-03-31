# -*- coding: utf-8 -*-
import sys
import openpyxl
import clr
pyt_path = r'C:\Python27\Lib\site-packages'
sys.path.append(pyt_path)
clr.AddReference("RevitAPI")
from Autodesk.Revit.DB import BuiltInCategory, FilteredElementCollector
clr.AddReference("RevitServices")
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager
doc = DocumentManager.Instance.CurrentDBDocument


def check_parameter(elem, param_name):
    document = elem.Document
    result = False
    param = elem.LookupParameter(param_name)
    if param:
        value = param.AsValueString()
        if value is None or value == "":
            result = True
    else:
        elem_type = document.GetElement(elem.GetTypeId())
        elem_type_param = elem_type.LookupParameter(param_name)
        if elem_type_param:
            value = elem_type_param.AsValueString()
            if value is None or value == "":
                result = True
        else:
            result = 'NP'
    return result


def write_data(data, document=doc):
    import os
    from io import open
    from datetime import datetime
    date = datetime.now()
    current_date = date.strftime("%Y%m%d_%H%M%S")
    homedir = os.path.expanduser('~') + f"\\Desktop\\{document.Title}_{current_date}.txt"
    with open(homedir, 'w', newline='') as file:
        for row in data:
            for r in row:
                file.write(r + '\n')
        file.close()
    return data


categories = {
    'Стены': BuiltInCategory.OST_Walls,
    'Перекрытия': BuiltInCategory.OST_Floors
}

xpath = r"C:\Users\s.shvydko\Desktop\тест1.xlsx"
workbook = openpyxl.load_workbook(xpath)
sheet = workbook['Лист1']
max_column = sheet.max_column
max_row = sheet.max_row

error = []
result = []
for y in range(2, max_column + 1):
    info = []
    cell_param_name = sheet.cell(row=1, column=y).value
    info.append(cell_param_name)
    for i in range(2, max_row + 1):
        cell_cat = sheet.cell(row=i, column=1).value
        cell_param_value = sheet.cell(row=i, column=y).value
        try:
            cat = categories[cell_cat]
            elems = FilteredElementCollector(doc).OfCategory(cat).WhereElementIsNotElementType().ToElements()
            for elem in elems:
                if cell_param_value is not None:
                    check_value = check_parameter(elem, cell_param_name)
                    if check_value is True:
                        info.append(f"Not filled | {elem.Name:^15} | {elem.Id}")
                    elif check_value == "NP":
                        info.append(f'No paramater | {elem.Name:^15} | {elem.Id}')
        except KeyError:
            error.append(cell_cat)
    if len(info) > 1:
        result.append(info)

if not error:
    error.append("No errors")

OUT = write_data(result), error
